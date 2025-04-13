import os
from groq import Groq
api_key = "replaced the api key"
client = Groq(api_key=api_key)
def summarize_email(email_body):
    try:
        chat_completion = client.chat.completions.create(
            model="meta-llama/llama-4-scout-17b-16e-instruct",
            messages=[
                {
                    "role": "system",
                    "content": (
                        "Summarize the text strictly in this format: 'Company: <Name> | Category: <Type> | Key Action: <Action>'. "
                        "Ensure the output is no more than 100 characters. "
                        "Ignore links but mention them in 'Key Action' if present. "
                        "If the email contains multiple items, mention it without listing them. "
                        "Example: 'Company: Jacobs | Category: Job Application | Key Action: Complete Submission, links included'."
                    ),
                },
                {
                    "role": "user",
                    "content": email_body,
                },
            ],
            temperature=1,
        max_completion_tokens=10,
        top_p=1,
        stream=False,
        stop=None,
        )

        return chat_completion.choices[0].message.content

    except Exception as e:
        print(f"Error summarizing email: {e}")
        # Return a default or empty value in case of error, or simply skip it
        return "Error summarizing this email"


import openpyxl

def process_excel_data(input_file, output_file, summarize_email):
    # Load the existing workbook
    wb = openpyxl.load_workbook(input_file)
    ws = wb.active

    # Create a new workbook for output
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.append(["Original Data", "Summarized Data"])

    # Process each cell in column A
    i=0
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
        for cell in row:
            i+=1
            try:
                summarized_text = summarize_email(cell.value)
                new_ws.append([cell.value, summarized_text])
                print("rows appended:",i)
            except Exception as e:
                print(f"Error processing email: {e}")
                new_ws.append([cell.value, "Error summarizing this email"])  # Handle error in a way you prefer

    # Save the processed data to a new file
    new_wb.save(output_file)
    print(f"Processed data saved to {output_file}")


input_file = "data.xlsx"
output_file = "output.xlsx"
process_excel_data(input_file, output_file, summarize_email)

